"""Historian export endpoint.

Operator request 2026-06-12: ship a real Excel export from the
Historian page with optional template placeholders so the customer
can pre-style a beautiful workbook (logo, fonts, colours, header
rows) and have the data just *land* in the right cells.

The endpoint accepts a list of pre-rendered rows (already filtered
client-side from the visible historianRows) plus a column spec, and
returns an .xlsx file the browser triggers download for.

Template mode:
  The operator uploads a `.xlsx` whose cells contain text-mode
  placeholders. Two kinds:
    - Single cell placeholder: any cell containing  e.g. "{{ts}}"
      OR "Report generated at {{ts}}" -- the substring is replaced
      with the corresponding value from the FIRST data row when the
      cell sits OUTSIDE the loop block, OR the current row when it
      sits INSIDE the loop block.
    - Loop block: a cell whose value is exactly "{{#each}}" marks
      the start of the data-loop row. The first cell on a later row
      with value "{{/each}}" marks the end. Every row BETWEEN those
      two markers is duplicated once per data row and placeholders
      inside the block resolve to that row's values.

  Anything else in the workbook (logos, headers, styled cells,
  formulas) is preserved.
"""

from __future__ import annotations

import base64
import io
import re
from datetime import datetime, timezone
from typing import Any, List

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field


router = APIRouter(prefix="/api/historian", tags=["historian"])


class ExportColumn(BaseModel):
    key: str
    label: str


class ExportXlsxRequest(BaseModel):
    rows: list[dict[str, Any]] = Field(default_factory=list)
    columns: list[ExportColumn] = Field(default_factory=list)
    template_xlsx_b64: str | None = None
    template_name: str | None = None


_PLACEHOLDER_PATTERN = re.compile(r"\{\{\s*([a-zA-Z0-9_./-]+)\s*\}\}")


def _resolve_placeholder(text: str, row: dict[str, Any]) -> str:
    """Replace every {{key}} in `text` with row[key] (stringified).

    Unknown keys are left untouched so the operator can spot typos in
    their template.
    """

    def _replace(match: re.Match[str]) -> str:
        key = match.group(1).strip()
        if key in row:
            v = row[key]
            return "" if v is None else str(v)
        # Allow case-insensitive lookup against labels (operator types
        # "{{Tag}}" against a column labelled "Tag").
        for k in row.keys():
            if str(k).lower() == key.lower():
                v = row[k]
                return "" if v is None else str(v)
        return match.group(0)

    return _PLACEHOLDER_PATTERN.sub(_replace, str(text or ""))


def _build_plain_xlsx(rows: list[dict[str, Any]], columns: list[ExportColumn]) -> bytes:
    """Header row + data rows, no styling. Used when no template was
    uploaded."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dep missing
        raise HTTPException(status_code=500, detail=f"openpyxl missing: {exc}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Historian"
    if columns:
        ws.append([c.label for c in columns])
    else:
        # Fallback: derive from first row keys.
        if rows:
            ws.append(list(rows[0].keys()))
            columns = [ExportColumn(key=k, label=k) for k in rows[0].keys()]
    for r in rows:
        ws.append([r.get(c.label, r.get(c.key, "")) for c in columns])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def _build_templated_xlsx(
    rows: list[dict[str, Any]],
    columns: list[ExportColumn],
    template_b64: str,
) -> bytes:
    """Load the operator's .xlsx, apply placeholder substitutions,
    repeat the loop block once per row."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dep missing
        raise HTTPException(status_code=500, detail=f"openpyxl missing: {exc}")
    try:
        raw = base64.b64decode(template_b64)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid template base64: {exc}")
    try:
        wb = load_workbook(io.BytesIO(raw))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not open template: {exc}")

    ws = wb.active

    # Identify the loop block. We look for the FIRST cell containing
    # exactly "{{#each}}" (trimmed) — its row is the loop-start row.
    # Then we find the FIRST cell at or after that row containing
    # "{{/each}}" — its row is the loop-end row. Inclusive on both
    # ends. If no markers found, every placeholder is single-row.
    loop_start: int | None = None
    loop_end: int | None = None
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            txt = v.strip()
            if loop_start is None and txt == "{{#each}}":
                loop_start = cell.row
            elif loop_start is not None and loop_end is None and txt == "{{/each}}":
                loop_end = cell.row
                break
        if loop_end is not None:
            break

    # Build the global row context — first data row's values, used to
    # fill placeholders OUTSIDE the loop. We also inject a few
    # synthetic placeholders so the operator can reference the
    # export wallclock and the number of rows in title rows etc.
    from datetime import datetime as _dt
    synthetic = {
        "ts": _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ts_iso": _dt.now().isoformat(timespec="seconds"),
        "row_count": str(len(rows)),
    }
    first_row = dict(rows[0] if rows else {})
    for k, v in synthetic.items():
        first_row.setdefault(k, v)

    def _apply_to_cell(cell, ctx: dict[str, Any]) -> None:
        v = cell.value
        if not isinstance(v, str):
            return
        new_v = _resolve_placeholder(v, ctx)
        # Don't overwrite when the cell was a pure marker token —
        # those should become empty so the report doesn't show the
        # literal "{{#each}}" text.
        if v.strip() in ("{{#each}}", "{{/each}}"):
            cell.value = ""
            return
        cell.value = new_v

    if loop_start is None or loop_end is None:
        # No loop: just substitute every placeholder using the first
        # row context. If the template has none, the operator gets
        # back exactly what they uploaded.
        for row in ws.iter_rows():
            for cell in row:
                _apply_to_cell(cell, first_row)
    else:
        # 1. Build the loop template as a list of "row patterns".
        loop_template_rows: list[list[Any]] = []
        for r_idx in range(loop_start, loop_end + 1):
            row_cells = []
            for cell in ws[r_idx]:
                row_cells.append(cell.value)
            loop_template_rows.append(row_cells)

        # 2. Substitute non-loop cells with the first row context.
        for row in ws.iter_rows():
            if loop_start <= row[0].row <= loop_end:
                continue
            for cell in row:
                _apply_to_cell(cell, first_row)

        # 3. Clear the loop block rows on the sheet — we'll re-emit
        #    them below.
        for r_idx in range(loop_start, loop_end + 1):
            for cell in ws[r_idx]:
                cell.value = None

        # 4. Insert one set of loop rows per data row, starting at
        #    loop_start. We append to the sheet at the bottom since
        #    inserting rows in-place would invalidate styling refs.
        #    Simpler: write to fresh rows starting at loop_start.
        write_row = loop_start
        from copy import copy as shallow_copy  # for cell styles
        for data_row in rows:
            # Merge synthetic context (ts, row_count) so loop cells
            # can also reference them — useful for footers like
            # "Generated {{ts}}" inside the loop block.
            merged_ctx = dict(data_row)
            for k, v in synthetic.items():
                merged_ctx.setdefault(k, v)
            for pattern_row in loop_template_rows:
                for c_idx, val in enumerate(pattern_row, start=1):
                    cell = ws.cell(row=write_row, column=c_idx)
                    if isinstance(val, str):
                        cell.value = _resolve_placeholder(val, merged_ctx)
                    else:
                        cell.value = val
                write_row += 1

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


@router.get("/export-xlsx/reference-template")
def reference_template() -> StreamingResponse:
    """Return a styled .xlsx reference template the operator can
    download, edit, and re-upload. Demonstrates every supported
    placeholder so the operator knows what they can drop in their
    own workbook.

    Layout:
      Row 1 — Title with the export wallclock placeholder.
      Row 3 — Header row with column labels (matches the default
              export columns).
      Row 5 — Loop start marker {{#each}} in column A.
      Row 6 — One row of placeholders {{Timestamp}} {{Tag}} {{Value}}
              {{Quality}} {{Device}} {{Gateway}}.
      Row 7 — Loop end marker {{/each}}.
      Row 9 — Helper text explaining how to use it.

    Operator uploads back into Export modal → backend repeats
    rows 5–7 for every data row, substituting placeholders.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"openpyxl missing: {exc}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reference"

    accent = PatternFill(start_color="14A89A", end_color="14A89A", fill_type="solid")
    light = PatternFill(start_color="EAF6F4", end_color="EAF6F4", fill_type="solid")
    head_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="0E1A2B")
    body_font = Font(name="Calibri", size=10, color="0E1A2B")
    thin = Side(border_style="thin", color="BBBBBB")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column widths
    widths = [22, 22, 14, 14, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w

    # Row 1 — title
    ws.cell(row=1, column=1, value="TrustNode Historian Export — generated {{ts}}").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Row 3 — column headers (the operator can rename them)
    headers = ["Timestamp", "Tag", "Value", "Quality", "Device", "Gateway"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = head_font
        c.fill = accent
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box

    # Row 5 — loop start marker (column A)
    ws.cell(row=5, column=1, value="{{#each}}").font = Font(italic=True, color="888888")

    # Row 6 — placeholder row inside the loop
    placeholders = ["{{Timestamp}}", "{{Tag}}", "{{Value}}", "{{Quality}}", "{{Device}}", "{{Gateway}}"]
    for i, p in enumerate(placeholders, start=1):
        c = ws.cell(row=6, column=i, value=p)
        c.font = body_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.fill = light
        c.border = box

    # Row 7 — loop end marker
    ws.cell(row=7, column=1, value="{{/each}}").font = Font(italic=True, color="888888")

    # Row 9+ — operator-facing instructions
    instructions = [
        "How to use this template:",
        "  1. Open it in Excel and style it however you like (colors, fonts, logos, merged cells, etc.).",
        "  2. The placeholders {{Timestamp}} / {{Tag}} / {{Value}} / {{Quality}} / {{Device}} / {{Gateway}}",
        "     are case-insensitive and match the column labels in the export modal. Add or remove",
        "     placeholders to suit your report.",
        "  3. The row between {{#each}} and {{/each}} is repeated once per data row.",
        "  4. Cells OUTSIDE the loop (titles, footers, summary rows) are substituted with the FIRST row's values.",
        "  5. {{ts}} is replaced with the export's wallclock time (the moment you click Export).",
        "  6. Save as .xlsx and upload via the Export modal's 'Excel template' field.",
    ]
    for i, text in enumerate(instructions):
        c = ws.cell(row=9 + i, column=1, value=text)
        c.font = body_font
        ws.merge_cells(start_row=9 + i, start_column=1, end_row=9 + i, end_column=6)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    headers_resp = {
        "Content-Disposition": 'attachment; filename="trustnode-historian-template-reference.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(iter([bio.read()]), media_type=headers_resp["Content-Type"], headers=headers_resp)


class ExportRangeRequest(BaseModel):
    """Server-side export: the caller says WHICH data, never carries it."""
    from_utc: str = ""
    to_utc: str = ""
    gateway: str = ""
    device: str = ""
    tag: str = ""
    edge_id: str = ""
    columns: List[str] = []
    max_rows: int = 0          # 0 = no ceiling other than the range itself
    chunk_rows: int = 5000     # rows per store query


_EXPORT_DEFAULT_COLUMNS = [
    "ts_utc", "gateway_name", "device_name", "plc_ip", "database_name",
    "tag_name", "value", "value_text", "data_type", "quality_label",
]


def _csv_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    if any(ch in text for ch in (",", '"', "\n", "\r")):
        return '"' + text.replace('"', '""') + '"'
    return text


def _stream_export_csv(req: "ExportRangeRequest"):
    """Yield CSV bytes, paging the historian in bounded chunks.

    Peak memory is one chunk, not one export — which is the whole point: the
    previous route held every row twice (once in the browser, once in the
    server's workbook) and fell over on a large range."""
    from app.state import app_store

    columns = [str(c).strip() for c in (req.columns or []) if str(c).strip()] or _EXPORT_DEFAULT_COLUMNS
    chunk = max(500, min(int(req.chunk_rows or 5000), 20000))
    ceiling = max(0, int(req.max_rows or 0))

    yield (",".join(_csv_cell(c) for c in columns) + "\r\n").encode("utf-8")

    sent = 0
    offset = 0
    while True:
        want = chunk if ceiling <= 0 else min(chunk, ceiling - sent)
        if want <= 0:
            break
        try:
            rows = app_store.get_historian_rows_range(
                from_utc=req.from_utc, to_utc=req.to_utc,
                limit=want, offset=offset, prefer_cloud_reads=False,
                gateway=req.gateway, device=req.device, tag=req.tag,
                edge_id=req.edge_id,
            ) or []
        except Exception as exc:
            # Surface the failure inside the file rather than truncating in
            # silence — an operator must never get a short export that looks whole.
            yield (f"# export interrupted after {sent} rows: "
                   f"{type(exc).__name__}: {exc}\r\n").encode("utf-8")
            return
        if not rows:
            break
        buf = []
        for row in rows:
            buf.append(",".join(_csv_cell(row.get(c)) for c in columns))
        yield ("\r\n".join(buf) + "\r\n").encode("utf-8")
        sent += len(rows)
        offset += len(rows)
        if len(rows) < want:
            break


@router.post("/export")
def export_range(payload: ExportRangeRequest) -> StreamingResponse:
    """Stream a historian export for a RANGE, without the rows ever passing
    through the browser (2026-08-22, item 12)."""
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"historian_{stamp}.csv"
    return StreamingResponse(
        _stream_export_csv(payload),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export-xlsx")
def export_xlsx(payload: ExportXlsxRequest) -> StreamingResponse:
    rows = payload.rows or []
    columns = payload.columns or []
    if payload.template_xlsx_b64:
        data = _build_templated_xlsx(rows, columns, payload.template_xlsx_b64)
    else:
        data = _build_plain_xlsx(rows, columns)
    name = payload.template_name or "historian.xlsx"
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name))[:80] or "historian.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{safe}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(iter([data]), media_type=headers["Content-Type"], headers=headers)
